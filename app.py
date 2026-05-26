"""
TikTok Comments Scraper
Powered by TikHub API — https://tikhub.io
"""

import streamlit as st
import requests
import time
import re
import io
from datetime import datetime
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TikTok Comments Scraper",
    page_icon="🎵",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Light theme ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
html, body, [data-testid="stAppViewContainer"] {
    background-color: #f7f8fc;
    color: #1a1a2e;
}
[data-testid="stSidebar"] {
    background-color: #ffffff;
    border-right: 1px solid #e8eaf0;
}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p { color: #1a1a2e !important; }

.metric-card {
    background: #ffffff;
    border: 1.5px solid #e8eaf0;
    border-top: 3px solid #e94560;
    border-radius: 10px;
    padding: 14px 16px;
    text-align: center;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
}
.metric-card .value { font-size: 1.8rem; font-weight: 700; color: #e94560; line-height: 1.2; }
.metric-card .label { font-size: 0.75rem; color: #6b7280; margin-top: 4px;
                      text-transform: uppercase; letter-spacing: 0.04em; }

.meta-card {
    background: #ffffff;
    border: 1.5px solid #e8eaf0;
    border-left: 4px solid #e94560;
    border-radius: 10px;
    padding: 18px 22px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    margin-bottom: 16px;
}
.meta-card h3 { margin: 0 0 4px 0; color: #1a1a2e; font-size: 1.1rem; }
.meta-row { display: flex; flex-wrap: wrap; gap: 20px; margin-top: 12px; }
.meta-item { display: flex; flex-direction: column; }
.meta-item .mi-label { font-size: 0.7rem; text-transform: uppercase;
                        letter-spacing: 0.06em; color: #9ca3af; margin-bottom: 2px; }
.meta-item .mi-value { font-size: 0.92rem; font-weight: 500; color: #1a1a2e; }

h1, h2, h3 { color: #1a1a2e !important; }

.stDownloadButton > button {
    border-radius: 8px !important;
    border: 1.5px solid #e94560 !important;
    color: #e94560 !important;
    background: #fff !important;
    font-weight: 500 !important;
}
.stDownloadButton > button:hover {
    background: #e94560 !important;
    color: #fff !important;
}
.stButton > button[kind="primary"] {
    background: #e94560 !important;
    border: none !important;
    border-radius: 8px !important;
    color: #fff !important;
    font-weight: 600 !important;
}
.stButton > button[kind="primary"]:hover { background: #c73652 !important; }

[data-testid="stTabs"] button { color: #6b7280 !important; font-weight: 500; }
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #e94560 !important;
    border-bottom-color: #e94560 !important;
}
.stProgress > div > div { background-color: #e94560 !important; }
[data-testid="stDataFrame"] { border: 1px solid #e8eaf0; border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
BASE_URL = "https://api.tikhub.io"


# ── Helpers ────────────────────────────────────────────────────────────────────
def make_headers(api_key: str) -> dict:
    return {"Authorization": f"Bearer {api_key}", "Accept": "application/json"}


def fmt_num(n) -> str:
    n = int(n or 0)
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1_000:
        return f"{n / 1_000:.1f}K"
    return str(n)


def fmt_time(unix_ts) -> str:
    if not unix_ts:
        return ""
    return datetime.fromtimestamp(int(unix_ts)).strftime("%Y-%m-%d %H:%M")


def fmt_duration(seconds) -> str:
    s = int(seconds or 0)
    return f"{s // 60}m {s % 60}s" if s >= 60 else f"{s}s"


# ── Video ID resolution ────────────────────────────────────────────────────────
def extract_video_id(value: str, headers: dict) -> str | None:
    value = value.strip()
    if value.isdigit():
        return value
    try:
        r = requests.get(
            f"{BASE_URL}/api/v1/tiktok/web/get_aweme_id",
            headers=headers, params={"url": value}, timeout=20,
        )
        if r.status_code == 200:
            vid = r.json().get("data", {}).get("aweme_id", "")
            if vid:
                return vid
    except Exception:
        pass
    m = re.search(r"/video/(\d+)", value)
    return m.group(1) if m else None


# ── Fetch video metadata ───────────────────────────────────────────────────────
def fetch_video_metadata(video_id: str, headers: dict) -> dict:
    """
    Tries the Web endpoint first, falls back to App V3.
    Returns a normalised metadata dict.
    """
    raw = None

    # Attempt 1 — Web API
    try:
        r = requests.get(
            f"{BASE_URL}/api/v1/tiktok/web/fetch_post_detail",
            headers=headers, params={"aweme_id": video_id}, timeout=30,
        )
        if r.status_code == 200:
            body = r.json().get("data") or {}
            raw  = body.get("aweme_detail") or body.get("item_info", {}).get("itemStruct") or body
    except Exception:
        pass

    # Attempt 2 — App V3
    if not raw:
        try:
            r = requests.get(
                f"{BASE_URL}/api/v1/tiktok/app/v3/fetch_one_video",
                headers=headers, params={"aweme_id": video_id}, timeout=30,
            )
            if r.status_code == 200:
                body = r.json().get("data") or {}
                raw  = body.get("aweme_detail") or body
        except Exception:
            pass

    if not raw:
        return {}

    author      = raw.get("author") or raw.get("authorMeta") or {}
    stats       = raw.get("statistics") or raw.get("stats") or {}
    music       = raw.get("music") or raw.get("musicMeta") or {}
    video_info  = raw.get("video") or {}
    created_ts  = raw.get("create_time") or raw.get("createTime") or 0

    return {
        "video_id":       video_id,
        "title":          raw.get("desc") or raw.get("text") or "",
        "author":         author.get("unique_id") or author.get("uniqueId") or author.get("nickname", ""),
        "author_name":    author.get("nickname") or "",
        "author_fans":    author.get("follower_count") or author.get("fans") or 0,
        "author_verified":author.get("custom_verify") or author.get("verified") or False,
        "created_at":     fmt_time(created_ts),
        "duration":       fmt_duration(video_info.get("duration") or raw.get("duration") or 0),
        "views":          stats.get("play_count")    or stats.get("playCount")    or 0,
        "likes":          stats.get("digg_count")    or stats.get("diggCount")    or 0,
        "comments_count": stats.get("comment_count") or stats.get("commentCount") or 0,
        "shares":         stats.get("share_count")   or stats.get("shareCount")   or 0,
        "bookmarks":      stats.get("collect_count") or stats.get("collectCount") or 0,
        "music_title":    music.get("title") or music.get("musicName") or "",
        "music_author":   music.get("author") or music.get("authorName") or "",
        "hashtags":       ", ".join(
            f"#{t.get('hashtagName') or t.get('title', '')}"
            for t in (raw.get("text_extra") or raw.get("challenges") or [])
            if (t.get("hashtagName") or t.get("title"))
        ),
        "url":            f"https://www.tiktok.com/@{author.get('unique_id','')}/video/{video_id}",
    }


# ── Fetch comments ─────────────────────────────────────────────────────────────
def fetch_comments(video_id, target, headers, progress_cb=None) -> list:
    comments, cursor = [], 0
    while len(comments) < target:
        count  = min(20, target - len(comments))
        params = {"aweme_id": video_id, "cursor": cursor,
                  "count": count, "current_region": ""}
        try:
            r = requests.get(
                f"{BASE_URL}/api/v1/tiktok/web/fetch_post_comment",
                headers=headers, params=params, timeout=30,
            )
        except Exception as exc:
            st.error(f"Request error: {exc}")
            break
        if r.status_code == 401:
            st.error("❌ Invalid API key (401).")
            break
        if r.status_code != 200:
            st.error(f"API error {r.status_code}: {r.text[:200]}")
            break
        data  = r.json().get("data") or {}
        batch = data.get("comments") or data.get("data") or []
        if not batch:
            break
        comments.extend(batch)
        cursor   = data.get("cursor", cursor + count)
        has_more = data.get("has_more", len(batch) == count)
        if progress_cb:
            progress_cb(min(len(comments) / target, 0.95))
        if not has_more:
            break
        if len(comments) < target:
            time.sleep(0.3)
    return comments[:target]


# ── Fetch replies (paginated) ─────────────────────────────────────────────────
def fetch_all_replies(video_id, parsed_comments, headers, progress_cb=None) -> None:
    """
    Per the TikTok Web API spec the reply endpoint accepts:
      item_id     — the video ID
      comment_id  — the parent comment ID
    Both aweme_id and item_id are tried for resilience.
    """
    eligible = [c for c in parsed_comments if c["reply_count"] > 0]
    total    = len(eligible)

    skipped = [c for c in eligible if not c["comment_id"]]
    if skipped:
        st.warning(
            f"⚠️ {len(skipped)} comment(s) have replies but no comment ID could be "
            f"resolved — their replies will be skipped. This usually means the comments "
            f"API returned an unexpected field name for the comment ID.",
            icon="⚠️",
        )

    eligible = [c for c in eligible if c["comment_id"]]

    for i, c in enumerate(eligible):
        all_replies = []
        cursor      = 0
        want        = max(c["reply_count"], 1)

        while len(all_replies) < want:
            # Try item_id first (per spec), fall back to aweme_id
            params = {
                "item_id":        video_id,
                "comment_id":     c["comment_id"],
                "cursor":         cursor,
                "count":          20,
                "current_region": "",
            }
            try:
                r = requests.get(
                    f"{BASE_URL}/api/v1/tiktok/web/fetch_post_comment_reply",
                    headers=headers, params=params, timeout=30,
                )
                # If item_id returns nothing, retry with aweme_id
                if r.status_code == 200:
                    data  = r.json().get("data") or {}
                    batch = data.get("comments") or []
                    if not batch:
                        # Retry with aweme_id param name
                        params2 = dict(params)
                        params2.pop("item_id")
                        params2["aweme_id"] = video_id
                        r2 = requests.get(
                            f"{BASE_URL}/api/v1/tiktok/web/fetch_post_comment_reply",
                            headers=headers, params=params2, timeout=30,
                        )
                        if r2.status_code == 200:
                            data  = r2.json().get("data") or {}
                            batch = data.get("comments") or []
                    if not batch:
                        break
                    all_replies.extend(batch)
                    cursor   = data.get("cursor", cursor + 20)
                    has_more = data.get("has_more", len(batch) == 20)
                    if not has_more:
                        break
                else:
                    break
                time.sleep(0.2)
            except Exception:
                break

        c["replies"] = all_replies
        if progress_cb and total > 0:
            progress_cb((i + 1) / total)
        time.sleep(0.2)


# ── Parse comment ──────────────────────────────────────────────────────────────
def parse_comment(c: dict, video_id: str) -> dict:
    user = c.get("user") or {}
    # TikTok Web API uses "cid"; App API may use "comment_id" — try all variants
    cid = (
        str(c.get("cid") or "").strip()
        or str(c.get("comment_id") or "").strip()
        or str(c.get("id") or "").strip()
    )
    return {
        "video_id":    video_id,
        "comment_id":  cid,
        "username":    user.get("unique_id") or user.get("nickname", ""),
        "text":        c.get("text") or c.get("comment_text", ""),
        "likes":       int(c.get("digg_count") or 0),
        "reply_count": int(c.get("reply_comment_total") or 0),
        "created_at":  fmt_time(c.get("create_time") or c.get("created_at")),
        "replies":     [],
    }


# ── XLSX export ────────────────────────────────────────────────────────────────
def build_xlsx(all_parsed: list, all_metadata: dict) -> bytes:
    wb = openpyxl.Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="E94560")
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    alt_fill = PatternFill("solid", fgColor="FFF5F7")

    def style_header(ws, headers, widths):
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    # Sheet 1 — Video Metadata
    ws0 = wb.active
    ws0.title = "Video Metadata"
    meta_headers = ["Video ID", "Title", "Author", "Display Name",
                    "Posted At", "Duration", "Views", "Likes",
                    "Comments", "Shares", "Bookmarks",
                    "Music", "Music Author", "Hashtags", "URL"]
    meta_widths  = [22, 50, 20, 25, 18, 10, 14, 14, 12, 12, 12, 30, 25, 40, 50]
    style_header(ws0, meta_headers, meta_widths)
    for ri, (vid, m) in enumerate(all_metadata.items(), 2):
        vals = [
            m.get("video_id",""), m.get("title",""), m.get("author",""),
            m.get("author_name",""), m.get("created_at",""), m.get("duration",""),
            m.get("views",0), m.get("likes",0), m.get("comments_count",0),
            m.get("shares",0), m.get("bookmarks",0),
            m.get("music_title",""), m.get("music_author",""),
            m.get("hashtags",""), m.get("url",""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws0.cell(row=ri, column=ci, value=val)
            cell.alignment = left if ci in (2, 14, 15) else center
            cell.border = thin
        ws0.row_dimensions[ri].height = 32

    # Sheet 2 — Comments + Replies as flat separate rows
    ws1 = wb.create_sheet("Comments")
    style_header(ws1,
        ["#", "Video ID", "Type", "Parent Username", "Username",
         "Text", "Likes", "Posted At"],
        [5, 22, 12, 22, 22, 65, 10, 18])

    reply_fill = PatternFill("solid", fgColor="FFF0F3")
    row_idx    = 2
    row_num    = 0

    for c in all_parsed:
        row_num += 1
        fill = alt_fill if row_num % 2 == 0 else None

        # Comment row
        row = [row_num, c["video_id"], "Comment", "",
               c["username"], c["text"], c["likes"], c["created_at"]]
        for ci, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=ci, value=val)
            cell.alignment = left if ci == 6 else center
            cell.border = thin
            if fill: cell.fill = fill
        ws1.row_dimensions[row_idx].height = 42
        row_idx += 1

        # Each reply as its own flat row
        for rep in c.get("replies") or []:
            row_num += 1
            ru   = rep.get("user") or {}
            rtxt = rep.get("text", "")
            rl   = int(rep.get("digg_count") or 0)
            rep_row = [row_num, c["video_id"], "Reply", c["username"],
                       ru.get("unique_id", ""), rtxt, rl, ""]
            for ci, val in enumerate(rep_row, 1):
                cell = ws1.cell(row=row_idx, column=ci, value=val)
                cell.fill      = reply_fill
                cell.border    = thin
                cell.alignment = left if ci == 6 else center
            ws1.row_dimensions[row_idx].height = 42
            row_idx += 1

    # Sheet 3 — Replies
    ws2 = wb.create_sheet("Replies")
    style_header(ws2,
        ["Comment ID", "Video ID", "Commenter", "Original Comment",
         "Reply By", "Reply Text", "Reply Likes"],
        [22, 20, 20, 45, 20, 55, 12])
    row_idx = 2
    for c in all_parsed:
        for rep in c.get("replies") or []:
            ru   = rep.get("user") or {}
            vals = [c["comment_id"], c["video_id"], c["username"], c["text"],
                    ru.get("unique_id", ""), rep.get("text", ""),
                    int(rep.get("digg_count") or 0)]
            for ci, val in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=ci, value=val)
                cell.alignment = left if ci in (4, 6) else center
                cell.border = thin
            ws2.row_dimensions[row_idx].height = 42
            row_idx += 1
    if row_idx == 2:
        ws2.cell(row=2, column=1, value="No replies fetched.").alignment = center

    # Sheet 4 — Summary
    ws3 = wb.create_sheet("Summary")
    style_header(ws3,
        ["Video ID", "Title", "Author", "Comments Scraped",
         "Total Likes", "Total Replies", "Top Commenter"],
        [22, 45, 20, 16, 14, 14, 25])
    for ri, vid in enumerate(dict.fromkeys(c["video_id"] for c in all_parsed), 2):
        grp   = [c for c in all_parsed if c["video_id"] == vid]
        m     = all_metadata.get(vid, {})
        top_u = Counter(c["username"] for c in grp).most_common(1)
        vals  = [vid, m.get("title","")[:80], m.get("author",""),
                 len(grp), sum(c["likes"] for c in grp),
                 sum(c["reply_count"] for c in grp),
                 top_u[0][0] if top_u else ""]
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.alignment = left if ci == 2 else center
            cell.border = thin
        ws3.row_dimensions[ri].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════════
#  UI
# ════════════════════════════════════════════════════════════════════════════════
st.title("🎵 TikTok Comments Scraper")
st.caption("Powered by [TikHub API](https://tikhub.io)")

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")
    api_key = st.text_input(
        "🔑 TikHub API Key",
        type="password",
        placeholder="Paste your API key here…",
        help="Get yours at tikhub.io → User Centre → API Token",
    )
    if api_key:
        st.success("API key set", icon="🔐")
    else:
        st.warning("Enter your API key to start.", icon="⚠️")

    st.divider()
    st.subheader("🎯 Scraping")
    num_comments = st.number_input(
        "Max comments per video",
        min_value=10,
        max_value=None,
        value=50,
        step=10,
        help="Enter any number — no upper limit. Large numbers will take longer.",
    )
    num_comments = int(num_comments)

    st.divider()
    st.subheader("🔍 Filters")
    min_likes      = st.number_input("Min likes on comment", min_value=0, value=0)
    keyword_filter = st.text_input("Keyword filter", placeholder="e.g. love")

    st.divider()
    st.caption("v1.0.0 · [TikHub Docs](https://docs.tikhub.io)")

# ── Input ──────────────────────────────────────────────────────────────────────
st.subheader("📋 Video IDs / URLs")
st.caption("Paste one TikTok video ID or URL per line. Mixed input is supported.")
raw_input = st.text_area(
    label="videos",
    label_visibility="collapsed",
    height=140,
    placeholder=(
        "7611351401800748308\n"
        "https://www.tiktok.com/@user/video/123456789\n"
        "https://vm.tiktok.com/ZMxxxxxxxx/"
    ),
)
run_btn = st.button("▶  Start Scraping", type="primary", use_container_width=True)

# ── Run ────────────────────────────────────────────────────────────────────────
if run_btn:
    if not api_key:
        st.error("Please enter your TikHub API key in the sidebar first.")
        st.stop()
    lines = [l.strip() for l in raw_input.strip().splitlines() if l.strip()]
    if not lines:
        st.error("Enter at least one video ID or URL above.")
        st.stop()

    headers:      dict = make_headers(api_key)
    all_parsed:   list = []
    all_metadata: dict = {}   # keyed by video_id
    ts_str             = datetime.now().strftime("%Y%m%d_%H%M%S")

    for idx, line in enumerate(lines):
        st.markdown(f"---\n### 🎬 Video {idx + 1} of {len(lines)}")
        st.code(line, language=None)

        # ── Resolve ID ────────────────────────────────────────────────────────
        with st.spinner("Resolving video ID…"):
            video_id = extract_video_id(line, headers)
        if not video_id:
            st.error("Could not resolve a video ID from the input above.")
            continue
        st.caption(f"Video ID: **{video_id}**")

        # ── Fetch metadata ────────────────────────────────────────────────────
        with st.spinner("Fetching video metadata…"):
            meta = fetch_video_metadata(video_id, headers)
        all_metadata[video_id] = meta

        if meta:
            verified = " ✅" if meta.get("author_verified") else ""
            st.markdown(
                f'<div class="meta-card">'
                f'<h3>{meta.get("title","(no title)") or "(no title)"}</h3>'
                f'<div class="meta-row">'
                f'<div class="meta-item"><span class="mi-label">Author</span>'
                f'<span class="mi-value">@{meta.get("author","")}{verified}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Display name</span>'
                f'<span class="mi-value">{meta.get("author_name","")}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Posted</span>'
                f'<span class="mi-value">{meta.get("created_at","")}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Duration</span>'
                f'<span class="mi-value">{meta.get("duration","")}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Views</span>'
                f'<span class="mi-value">{fmt_num(meta.get("views",0))}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Likes</span>'
                f'<span class="mi-value">{fmt_num(meta.get("likes",0))}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Comments</span>'
                f'<span class="mi-value">{fmt_num(meta.get("comments_count",0))}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Shares</span>'
                f'<span class="mi-value">{fmt_num(meta.get("shares",0))}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Bookmarks</span>'
                f'<span class="mi-value">{fmt_num(meta.get("bookmarks",0))}</span></div>'
                f'<div class="meta-item"><span class="mi-label">Music</span>'
                f'<span class="mi-value">{meta.get("music_title","") or "—"}</span></div>'
                f'</div>'
                + (f'<div style="margin-top:10px;font-size:0.82rem;color:#6b7280;">'
                   f'🏷 {meta.get("hashtags","")}</div>' if meta.get("hashtags") else "")
                + f'</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info("Metadata could not be fetched for this video.")

        # ── Fetch comments ────────────────────────────────────────────────────
        st.markdown("**Step 1 of 2 — Fetching comments**")
        prog1 = st.progress(0.0)
        raw   = fetch_comments(video_id, num_comments, headers,
                               progress_cb=lambda p: prog1.progress(p))
        prog1.progress(1.0, text=f"✅ {len(raw)} comments fetched")

        parsed = [parse_comment(c, video_id) for c in raw]

        # Filters
        filtered = parsed
        if min_likes > 0:
            filtered = [c for c in filtered if c["likes"] >= min_likes]
        if keyword_filter.strip():
            kw = keyword_filter.strip().lower()
            filtered = [c for c in filtered if kw in c["text"].lower()]

        # ── Fetch replies ─────────────────────────────────────────────────────
        eligible = [c for c in filtered if c["reply_count"] > 0]
        if eligible:
            st.markdown(f"**Step 2 of 2 — Fetching replies** ({len(eligible)} comment(s) have replies)")
            prog2 = st.progress(0.0)
            fetch_all_replies(video_id, filtered, headers,
                              progress_cb=lambda p: prog2.progress(p))
            total_replies = sum(len(c["replies"]) for c in filtered)
            prog2.progress(1.0, text=f"✅ {total_replies} replies fetched")
        else:
            st.markdown("**Step 2 of 2 — No replies found for these comments**")

        all_parsed.extend(filtered)

        # ── Metrics ───────────────────────────────────────────────────────────
        n           = max(len(filtered), 1)
        total_likes = sum(c["likes"] for c in filtered)
        tot_rep     = sum(len(c["replies"]) for c in filtered)

        st.write("")
        m1, m2, m3 = st.columns(3)
        for col, val, label in [
            (m1, len(filtered),        "Comments Scraped"),
            (m2, fmt_num(total_likes), "Total ❤ on Comments"),
            (m3, tot_rep,              "Replies Fetched"),
        ]:
            col.markdown(
                f'<div class="metric-card">'
                f'<div class="value">{val}</div>'
                f'<div class="label">{label}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        st.write("")

        # ── Tabs ──────────────────────────────────────────────────────────────
        tab1, tab2 = st.tabs(["💬 Comments", "🔑 Top Keywords"])

        with tab1:
            df = pd.DataFrame([
                {
                    "#":         i + 1,
                    "Username":  c["username"],
                    "Comment":   c["text"],
                    "Likes":     c["likes"],
                    "Replies":   c["reply_count"],
                    "Posted At": c["created_at"],
                }
                for i, c in enumerate(filtered)
            ])
            st.dataframe(
                df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "#":         st.column_config.NumberColumn(width="small"),
                    "Username":  st.column_config.TextColumn(width="medium"),
                    "Comment":   st.column_config.TextColumn(width="large"),
                    "Likes":     st.column_config.NumberColumn(width="small", format="%d"),
                    "Replies":   st.column_config.NumberColumn(width="small", format="%d"),
                    "Posted At": st.column_config.TextColumn(width="medium"),
                },
            )

        with tab2:
            from collections import Counter as _Counter
            import re as _re
            try:
                from nltk.corpus import stopwords as _sw
                stop_words = set(_sw.words("english"))
            except Exception:
                stop_words = set()
            words = []
            for c in filtered:
                for w in _re.findall(r"[a-zA-Z]{3,}", c["text"].lower()):
                    if w not in stop_words:
                        words.append(w)
            kws = _Counter(words).most_common(20)
            if kws:
                kw_cols = st.columns(4)
                for i, (word, freq) in enumerate(kws):
                    kw_cols[i % 4].metric(word, freq)
            else:
                st.info("Not enough text to extract keywords.")

    # ── Global export ──────────────────────────────────────────────────────────
    if all_parsed:
        st.markdown("---")
        st.subheader("📤 Export")
        with st.spinner("Building XLSX…"):
            xlsx_bytes = build_xlsx(all_parsed, all_metadata)
        fname_xlsx = f"tiktok_comments_{ts_str}.xlsx"
        st.download_button(
            label=f"⬇  Download XLSX — {len(all_parsed)} comments · 4 sheets",
            data=xlsx_bytes,
            file_name=fname_xlsx,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption(f"`{fname_xlsx}` · **Video Metadata** · Comments · Replies · Summary")
